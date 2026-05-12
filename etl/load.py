#!/usr/bin/env python3
"""
Thazin & Cherry historical ETL.

Reads the 33 Excel source files and emits batched SQL files into etl/sql/
that can be executed via the Supabase MCP `execute_sql` tool (or psql).

The SQL is idempotent where possible: levels/discount_types/chart_of_accounts
were seeded by the migration, sections/students/etc. use ON CONFLICT to dedupe.

Run:  python3 etl/load.py
"""
from __future__ import annotations
import os, re, sys, hashlib, datetime
from pathlib import Path
from collections import defaultdict
from typing import Any, Iterable
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
FIN  = ROOT / "2026 Thazin&Cherry Finance"
STU  = ROOT / "2026 Thazin&Cherry student_s list"
OUT  = ROOT / "etl" / "sql"
OUT.mkdir(parents=True, exist_ok=True)

# ─── Maps ──────────────────────────────────────────────────────────────────
LEVEL_CODE = {
    "EARLY CHILDHOOD": "EARLY_CHILDHOOD",
    "EARLYCHILDHOOD": "EARLY_CHILDHOOD",
    "EC": "EARLY_CHILDHOOD",
    "NURSERY": "NURSERY",
    "PRE-STARTER": "PRE_STARTER",
    "PRE STARTER": "PRE_STARTER",
    "PRESTARTER": "PRE_STARTER",
    "PRE-STARTERS": "PRE_STARTER",
    "STARTER": "STARTER",
    "STARTERS": "STARTER",
    "MOVERS": "MOVERS",
    "MOVER": "MOVERS",
    "FLYERS": "FLYERS",
    "FLYER": "FLYERS",
    "KEY": "KEY",
    "PET": "PET",
    "FCE": "FCE",
    "CAE": "CAE",
}

PER_LEVEL_FILES = [
    ("EARLY_CHILDHOOD", STU / "2026 T& C Students_ list & Attendance Record/Early Childhood 2026 Ss Lists and Attendance_.xlsx"),
    ("NURSERY",         STU / "2026 T& C Students_ list & Attendance Record/Nursery  2026 Ss Lists and Attendance_.xlsx"),
    ("PRE_STARTER",     STU / "2026 T& C Students_ list & Attendance Record/Pre-Starters 2026 Ss Lists and Attendance_.xlsx"),
    ("STARTER",         STU / "2026 T& C Students_ list & Attendance Record/Starters 2026 Ss Lists and Attendance_.xlsx"),
    ("MOVERS",          STU / "2026 T& C Students_ list & Attendance Record/Movers 2026 Ss Lists and Attendance_.xlsx"),
    ("FLYERS",          STU / "2026 T& C Students_ list & Attendance Record/Flyers 2026 Ss Lists and Attendance_.xlsx"),
    ("KEY",             STU / "2026 T& C Students_ list & Attendance Record/KEY 2026 Ss Lists and Attendance_.xlsx"),
    ("PET",             STU / "2026 T& C Students_ list & Attendance Record/PET 2026 Ss Lists and Attendance_.xlsx"),
    ("FCE",             STU / "2026 T& C Students_ list & Attendance Record/FCE 2026 Ss Lists and Attendance_.xlsx"),
    ("CAE",             STU / "2026 T& C Students_ list & Attendance Record/CAE 2026 Ss Lists and Attendance_.xlsx"),
]
COMBINED_FILES = [
    ("EARLY_CHILDHOOD", STU / "2026 T&C Early Childhood Students_ list & Attendance List_.xlsx"),
    ("NURSERY",         STU / "2026 T&C Nursery Students_ list & Attendance List_.xlsx"),
]

KPAY_FILES = [
    ("2026-01", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Jan  Daily K Pay Transcation.xlsx"),
    ("2026-02", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Feb  Daily K Pay Transcation.xlsx"),
    ("2026-03", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 March  Daily K Pay Transcation.xlsx"),
    ("2026-04", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 April  Daily K Pay Transcation.xlsx"),
]

INCOME_STATEMENT_FILES = [
    ("2026-01", FIN / "All Months Income Statement/January Income Statement.xlsx"),
    ("2026-02", FIN / "All Months Income Statement/February Income Statement.xlsx"),
    ("2026-03", FIN / "All Months Income Statement/March Income Statement.xlsx"),
    ("2026-04", FIN / "All Months Income Statement/April Income Statement.xlsx"),
]

EVENT_FILES = [
    ("Thingyan Festival 2026",  FIN / "2026 T & C Event Budget /2026 Thingyan Festival Cost.xlsx", datetime.date(2026, 4, 13)),
    ("Awarding Ceremony 2026",  FIN / "2026 T & C Event Budget /2026 Awarding Budget_.xlsx",       datetime.date(2026, 5, 1)),
]

CONFIRM_PRICE = FIN / "T&C All Program Opening Cost/2026 ESL Confirm Price.xlsx"

# ─── Helpers ───────────────────────────────────────────────────────────────
def q(v: Any) -> str:
    """SQL-quote a Python value safely."""
    if v is None or v == "":
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int,)):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN
            return "NULL"
        if v.is_integer():
            return str(int(v))
        return repr(v)
    if isinstance(v, datetime.datetime):
        return f"'{v.isoformat(sep=' ')}'"
    if isinstance(v, datetime.date):
        return f"'{v.isoformat()}'"
    s = str(v).replace("'", "''")
    return f"'{s}'"

def clean_phone(s: str | None) -> tuple[str | None, str | None]:
    if not s: return (None, None)
    s = re.sub(r"[\s\n]+", " ", str(s)).strip()
    parts = re.split(r"[/,;]+", s)
    parts = [p.strip() for p in parts if p.strip()]
    return (parts[0] if parts else None, parts[1] if len(parts) > 1 else None)

def normalize_status(s: Any) -> str:
    if s is None: return "Active"
    t = str(s).strip().lower()
    if "active" in t: return "Active"
    if "break"  in t: return "Break"
    if "left"   in t: return "Left"
    return "Active"

def parse_section_label(sheet_name: str, level_code: str) -> tuple[str, bool]:
    """Return (time_slot, is_online) parsed from the sheet name."""
    s = sheet_name.strip()
    # Strip leading level word(s)
    s = re.sub(r"(?i)^(early childhood|nursery|pre[\s-]?starters?|starters?|movers?|flyers?|key|pet|fce|cae)\b", "", s).strip()
    is_online = "online" in s.lower()
    s = re.sub(r"(?i)online", "", s).strip()
    m = re.search(r"\(?\s*([\d.:\-\s]+)\s*\)?", s)
    slot = m.group(1).strip() if m else (s or "default")
    # Normalize "745-945" → "7:45-9:45", "315-515" → "3:15-5:15", "10-12" stays.
    slot = slot.replace(" ", "")
    m2 = re.fullmatch(r"(\d{1,2})(\d{2})-(\d{1,2})(\d{2})", slot)
    if m2:
        slot = f"{int(m2.group(1))}:{m2.group(2)}-{int(m2.group(3))}:{m2.group(4)}"
    return (slot, is_online)

def extract_teacher_names(raw: Any) -> list[tuple[str, str | None]]:
    """Return list of (name, weekday_pattern). Handles co-teach 'Tr A(Sat)+Tr B(Sun)'."""
    if not raw: return []
    s = str(raw).replace("\n", " ").strip()
    parts = re.split(r"\s*\+\s*|\s*&\s*", s)
    out = []
    for p in parts:
        p = p.strip()
        m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", p)
        if m:
            out.append((m.group(1).strip(), m.group(2).strip()))
        elif p:
            out.append((p, None))
    return out

def to_int_money(v: Any) -> int | None:
    if v is None or v == "": return None
    if isinstance(v, (int, float)):
        return int(v) if v == v else None  # NaN guard
    s = str(v).strip().replace(",", "").replace("MMK", "").replace("ks", "").strip()
    if not s: return None
    try: return int(float(s))
    except ValueError: return None

def parse_date(v: Any) -> datetime.date | None:
    if v is None or v == "": return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def parse_datetime(v: Any) -> datetime.datetime | None:
    if v is None or v == "": return None
    if isinstance(v, datetime.datetime): return v
    if isinstance(v, datetime.date): return datetime.datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y"):
        try: return datetime.datetime.strptime(s, fmt)
        except ValueError: continue
    return None

def write_batched(path: Path, header: str, body_lines: list[str], footer: str = "", per_batch: int = 400):
    """Write SQL file with batched multi-row INSERTs."""
    with path.open("w") as f:
        f.write("BEGIN;\n")
        f.write(header.rstrip() + "\n\n" if header else "")
        if not body_lines:
            f.write("-- (no rows)\n")
        else:
            i = 0
            while i < len(body_lines):
                chunk = body_lines[i:i+per_batch]
                f.write(",\n".join(chunk) + ";\n\n")
                i += per_batch
        if footer:
            f.write(footer.rstrip() + "\n")
        f.write("COMMIT;\n")

# ─── Stage tables we'll use heavily ────────────────────────────────────────
# We can't generate IDs client-side, so we use natural keys via temp tables
# + INSERT…SELECT joins. Single SQL file per concept, all wrapped in BEGIN/COMMIT.

# ─── 1. Fee schedule from "2026 ESL Confirm Price.xlsx" ────────────────────
def gen_fee_schedule() -> list[str]:
    """Returns a list of complete SQL statements (semicolon-terminated)."""
    if not CONFIRM_PRICE.exists():
        return ["-- Confirm Price file missing"]
    wb = openpyxl.load_workbook(CONFIRM_PRICE, data_only=True)
    ws = wb["Summary"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    out.append("-- Fee schedule from 2026 ESL Confirm Price.xlsx Summary sheet")
    out.append("delete from fee_schedule where effective_from = '2026-01-01';")
    inserts: list[str] = []
    seen_codes: set[str] = set()
    # Header row at index 2 (row 3): Class Name, Class Fee, Textbook, T-Shirt, ID Card, Utilities, Discount, old total, New
    for r in rows[3:]:
        if not r or not r[0]: continue
        name = str(r[0]).strip()
        if not name or name.lower() in ("class name","total"): continue
        code = LEVEL_CODE.get(name.upper())
        if not code:
            code = LEVEL_CODE.get(name.upper().replace("-","").replace(" ",""))
        if not code: continue
        if code in seen_codes:        # keep only first row per level (main ESL grid)
            continue
        class_fee = to_int_money(r[1]) or 0
        textbook  = to_int_money(r[2]) or 0
        tshirt    = to_int_money(r[3]) or 0
        idcard    = to_int_money(r[4]) or 0
        if class_fee == 0 and textbook == 0:
            continue
        seen_codes.add(code)
        inserts.append(
            f"  ((select id from levels where code={q(code)}), '2026-01-01', null, "
            f"{class_fee}, {textbook}, {tshirt}, {idcard}, 0, 0)"
        )
    if inserts:
        out.append("insert into fee_schedule "
                   "(level_id, effective_from, effective_to, class_fee, textbook_fee, tshirt_fee, id_card_fee, guide_fee, default_discount) values\n"
                   + ",\n".join(inserts) + ";")
    wb.close()
    return out

# ─── 2. Sections + Teachers + Students + Enrolments + Sessions + Attendance ─
def parse_attendance_workbook(level_code: str, fp: Path):
    """Yield (section_natural_key, teachers, students, sessions, marks)."""
    wb = openpyxl.load_workbook(fp, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in ("total count","summary"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4: continue

        time_slot, is_online = parse_section_label(sheet_name, level_code)
        section_key = f"{level_code}|{time_slot}|{is_online}"

        # teacher row (R2 = index 1), look for the teacher names cell (col index 2 onwards)
        teacher_raw = None
        if len(rows) > 1:
            for v in rows[1][1:6]:
                if v and str(v).strip().lower().startswith("tr"):
                    teacher_raw = v; break
            if teacher_raw is None and rows[1][2]:
                teacher_raw = rows[1][2]
        teachers = extract_teacher_names(teacher_raw)

        # date headers from row 2 (index 1), starting where datetimes appear
        header_row = rows[1] if len(rows) > 1 else ()
        date_cols: list[tuple[int, datetime.date]] = []
        for i, v in enumerate(header_row):
            d = parse_date(v) if isinstance(v, (datetime.datetime, datetime.date, str)) else None
            if d: date_cols.append((i, d))

        # Student rows from R4 onward (index 3+)
        # Column layout per R3: 0=No., 1=MM, 2=EN, 3=Viber, 4=Phone, 5=Status, 7=Sat, 8=Sun, ...
        students = []
        marks = []
        for r in rows[3:]:
            if not r or not isinstance(r, tuple): continue
            no = r[0] if len(r)>0 else None
            mm = r[1] if len(r)>1 else None
            en = r[2] if len(r)>2 else None
            viber = r[3] if len(r)>3 else None
            phone = r[4] if len(r)>4 else None
            status = r[5] if len(r)>5 else None
            if not en and not mm: continue
            if isinstance(no, str) and no.strip().lower() in ("no.","total","count"): continue
            student_key = f"{(en or '').strip()}|{(mm or '').strip()}|{(phone or '').strip()[:20]}"
            students.append({
                "key": student_key,
                "english_name": (str(en).strip() if en else None),
                "myanmar_name": (str(mm).strip() if mm else None),
                "phone_raw": phone,
                "viber": (str(viber).strip() if viber else None),
                "status": normalize_status(status),
            })
            for col_idx, dt in date_cols:
                if col_idx >= len(r): continue
                cell = r[col_idx]
                if cell is True:
                    marks.append((student_key, dt, "Present"))
                elif cell is False:
                    marks.append((student_key, dt, "Absent"))
                # ignore None / strings

        sessions = sorted({d for _, d in date_cols})
        yield {
            "section_key": section_key,
            "level_code": level_code,
            "time_slot": time_slot,
            "is_online": is_online,
            "teachers": teachers,
            "students": students,
            "sessions": sessions,
            "marks": marks,
        }
    wb.close()

def gen_students_attendance() -> dict[str, list[str]]:
    """Build per-table SQL files. Returns dict of file_name → SQL lines."""
    sections_seen: dict[str, dict] = {}
    teachers_seen: dict[str, str] = {}    # short_name → full_name
    students_seen: dict[str, dict] = {}   # student_key → record (best version)
    enrolments_seen: list[tuple[str,str,datetime.date,str]] = []  # student_key, section_key, start, status
    sessions_seen: set[tuple[str, datetime.date]] = set()
    marks_seen: list[tuple[str,str,datetime.date,str]] = []        # student_key, section_key, date, status
    section_teacher_links: list[tuple[str,str,str|None]] = []      # section_key, teacher_short, weekday

    files = []
    for code, fp in PER_LEVEL_FILES + COMBINED_FILES:
        if fp.exists(): files.append((code, fp))
        else: print(f"  ! missing {fp}", file=sys.stderr)

    for code, fp in files:
        for sec in parse_attendance_workbook(code, fp):
            # Drop sections without a parseable time slot (combined files etc.)
            if not sec["time_slot"] or sec["time_slot"].strip() in ("", "default"):
                continue
            sections_seen[sec["section_key"]] = {
                "level_code": sec["level_code"],
                "time_slot":  sec["time_slot"],
                "is_online":  sec["is_online"],
            }
            for tname, weekday in sec["teachers"]:
                short = tname.strip()
                if not short or short.lower().startswith("e.g."): continue
                teachers_seen[short] = short
                section_teacher_links.append((sec["section_key"], short, weekday))
            for st in sec["students"]:
                if not st["english_name"] and not st["myanmar_name"]: continue
                prev = students_seen.get(st["key"])
                if prev is None or (prev["status"] == "Left" and st["status"] != "Left"):
                    students_seen[st["key"]] = st
                start = sec["sessions"][0] if sec["sessions"] else datetime.date(2026,1,1)
                enrolments_seen.append((st["key"], sec["section_key"], start, st["status"]))
            for d in sec["sessions"]:
                sessions_seen.add((sec["section_key"], d))
            for sk, d, status in sec["marks"]:
                marks_seen.append((sk, sec["section_key"], d, status))

    out: dict[str, list[str]] = {}

    # 1. Teachers
    teacher_rows = [
        f"  ({q(short)}, {q(short)})"
        for short in sorted(teachers_seen.keys())
    ]
    out["02_teachers.sql"] = [
        "insert into teachers (full_name, short_name) values\n"
        + ",\n".join(teacher_rows) + "\non conflict do nothing;"
    ] if teacher_rows else ["-- no teachers"]

    # 2. Sections
    section_rows = []
    for k, sec in sections_seen.items():
        section_rows.append(
            f"  ((select id from levels where code={q(sec['level_code'])}), {q(sec['time_slot'])}, {str(sec['is_online']).upper()})"
        )
    out["03_sections.sql"] = [
        "insert into sections (level_id, time_slot, is_online) values\n"
        + ",\n".join(section_rows) + "\non conflict (level_id, time_slot, is_online) do nothing;"
    ] if section_rows else ["-- no sections"]

    # 3. Section ↔ teacher
    st_rows = []
    seen_st = set()
    for sk, tn, weekday in section_teacher_links:
        key = (sk, tn)
        if key in seen_st: continue
        seen_st.add(key)
        sec = sections_seen[sk]
        st_rows.append(
            "  ("
            f"(select id from sections where level_id=(select id from levels where code={q(sec['level_code'])}) and time_slot={q(sec['time_slot'])} and is_online={str(sec['is_online']).upper()}),"
            f"(select id from teachers where short_name={q(tn)} limit 1),"
            f"{q(weekday)}"
            ")"
        )
    out["04_section_teachers.sql"] = [
        "insert into section_teachers (section_id, teacher_id, weekday_pattern) values\n"
        + ",\n".join(st_rows) + "\non conflict do nothing;"
    ] if st_rows else ["-- no section teachers"]

    # 4. Guardians + Students
    # We use a deterministic external_id so we can join later.
    student_inserts = []
    for key, st in sorted(students_seen.items()):
        ext = "S:" + hashlib.sha1(key.encode()).hexdigest()[:14]
        st["external_id"] = ext
        phone1, phone2 = clean_phone(st["phone_raw"])
        # Insert guardian inline via CTE pattern (one per student)
        student_inserts.append((ext, st, phone1, phone2))

    # Strategy: stage guardians with returning, then students. Use one combined CTE per chunk.
    # Simpler: for each student, two statements; batch them.
    g_rows = []
    s_rows = []
    for ext, st, p1, p2 in student_inserts:
        g_rows.append(
            f"  ({q(ext)}, {q(p1)}, {q(p2)}, {q(st['viber'])})"
        )
        s_rows.append(
            f"  ({q(ext)}, "
            f"(select id from guardians where notes={q(ext)}), "
            f"{q(st['english_name'])}, {q(st['myanmar_name'])}, {q(st['status'])})"
        )
    out["05_guardians.sql"] = [
        "-- Use guardians.notes as a temporary external_id link\n"
        "insert into guardians (notes, phone_primary, phone_secondary, viber_number) values\n"
        + ",\n".join(g_rows) + "\non conflict do nothing;"
    ] if g_rows else ["-- no guardians"]

    out["06_students.sql"] = [
        "insert into students (external_id, guardian_id, english_name, myanmar_name, current_status) values\n"
        + ",\n".join(s_rows) + "\non conflict (external_id) do nothing;"
    ] if s_rows else ["-- no students"]

    # 5. Enrolments
    e_rows = []
    seen_e = set()
    for sk_student, sk_section, start, status in enrolments_seen:
        key = (sk_student, sk_section, start)
        if key in seen_e: continue
        seen_e.add(key)
        sec = sections_seen[sk_section]
        ext = "S:" + hashlib.sha1(sk_student.encode()).hexdigest()[:14]
        e_rows.append(
            "  ("
            f"(select id from students where external_id={q(ext)}),"
            f"(select id from sections where level_id=(select id from levels where code={q(sec['level_code'])}) and time_slot={q(sec['time_slot'])} and is_online={str(sec['is_online']).upper()}),"
            f"{q(start)}, {q(status)}"
            ")"
        )
    out["07_enrolments.sql"] = [
        "insert into enrolments (student_id, section_id, start_date, status) values\n"
        + ",\n".join(e_rows) + "\non conflict (student_id, section_id, start_date) do nothing;"
    ] if e_rows else ["-- no enrolments"]

    # 6. Class sessions
    cs_rows = []
    for sk, d in sorted(sessions_seen):
        sec = sections_seen[sk]
        cs_rows.append(
            "  ("
            f"(select id from sections where level_id=(select id from levels where code={q(sec['level_code'])}) and time_slot={q(sec['time_slot'])} and is_online={str(sec['is_online']).upper()}),"
            f"{q(d)}"
            ")"
        )
    out["08_class_sessions.sql"] = [
        "insert into class_sessions (section_id, session_date) values\n"
        + ",\n".join(cs_rows) + "\non conflict (section_id, session_date) do nothing;"
    ] if cs_rows else ["-- no sessions"]

    # 7. Attendance marks — chunk into multiple INSERT statements (each ≤ ~400 rows)
    am_rows = []
    seen_am = set()
    for sk_student, sk_section, d, status in marks_seen:
        key = (sk_student, sk_section, d)
        if key in seen_am: continue
        seen_am.add(key)
        sec = sections_seen[sk_section]
        ext = "S:" + hashlib.sha1(sk_student.encode()).hexdigest()[:14]
        am_rows.append(
            "  ("
            f"(select id from class_sessions where section_id=(select id from sections where level_id=(select id from levels where code={q(sec['level_code'])}) and time_slot={q(sec['time_slot'])} and is_online={str(sec['is_online']).upper()}) and session_date={q(d)}),"
            f"(select id from students where external_id={q(ext)}),"
            f"{q(status)}"
            ")"
        )

    # Chunk into per-file SQL of 400 rows each so each execute_sql payload stays small.
    BATCH = 400
    am_files: list[str] = []
    if am_rows:
        for i in range(0, len(am_rows), BATCH):
            chunk = am_rows[i:i+BATCH]
            stmt = (
                "insert into attendance_marks (session_id, student_id, status) values\n"
                + ",\n".join(chunk)
                + "\non conflict do nothing;"
            )
            am_files.append(stmt)
    else:
        am_files = ["-- no attendance"]
    # We split into multiple files: 09_attendance_marks_001.sql, _002.sql ...
    out["__ATTENDANCE_CHUNKS__"] = am_files  # special key, handled in main()

    return out

# ─── 3. KPay raw transactions ──────────────────────────────────────────────
def gen_kpay_transactions() -> list[str]:
    rows = []
    for tag, fp in KPAY_FILES:
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        if "K Pay Daily Transcation" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["K Pay Daily Transcation"]
        for r in ws.iter_rows(min_row=5, values_only=True):
            if not r or not r[0]: continue
            ts = parse_datetime(r[0])
            if not ts: continue
            sender_mm = (str(r[1]).strip() if r[1] else None)
            sender_en = (str(r[2]).strip() if r[2] else None)
            cls = (str(r[3]).strip() if r[3] else None)
            ptype = (str(r[4]).strip() if r[4] else None)
            amt = to_int_money(r[5] if len(r) > 5 else None)
            if amt is None or amt <= 0: continue
            rows.append(
                f"  ({q(ts)}, {q(sender_mm)}, {q(sender_en)}, {q(cls)}, {q(ptype)}, {amt}, {q(fp.name)})"
            )
        wb.close()
    if not rows: return ["-- no kpay rows"]
    return [
        "insert into kpay_transactions (txn_at, sender_mm, sender_en, raw_class, payment_type, amount, source_file) values\n"
        + ",\n".join(rows) + ";"
    ]

# ─── 4. Ledger entries from Income Statement files ─────────────────────────
def lookup_account(name: Any) -> str | None:
    if not name: return None
    s = str(name).strip().lower()
    canon_map = {
        "esl class fee": "ESL Class Fee",
        "summer program fee": "Summer Program Fee",
        "other income": "Other Income",
        "esl teacher salary": "ESL Teacher Salary",
        "admin teacher salary": "Admin Teacher Salary",
        "admin salary": "Admin Salary",
        "teaching supply": "Teaching Supply",
        "office expense": "Office Expense",
        "monthly operating expense": "Monthly Operating Expense",
        "monthly operation expense": "Monthly Operating Expense",
        "monthly opreation expense": "Monthly Operating Expense",
        "initial capital & major operating costs": "Initial Capital & Major Operating Costs",
        "one-time minor expense": "One-time Minor Expense",
        "one time minor expense": "One-time Minor Expense",
        "one-time capital & large operational expense": "One-time Capital & Large Operational Expense",
        "internet & communication expense": "Internet & Communication Expense",
        "drinking water": "Drinking Water",
        "delivery & transportation": "Delivery & Transportation",
        "office stationery": "Office Stationery",
        "government tax": "Government Tax",
        "event": "Event",
        "personal expense": "Personal Expense",
        "special case": "Special Case",
        "other expense": "Other Expense",
    }
    for k, v in canon_map.items():
        if k in s: return v
    return None

def gen_ledger_entries() -> list[str]:
    rows = []
    for tag, fp in INCOME_STATEMENT_FILES:
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        for src, sn in (("GeneralExpense", "General Expense"), ("OfficeExpense", "Office Expense")):
            if sn not in wb.sheetnames: continue
            ws = wb[sn]
            for r in ws.iter_rows(min_row=5, values_only=True):
                if not r: continue
                first = r[0]
                # Skip blank / header / summary rows
                if first is None and (len(r) < 4 or r[1] is None): continue
                d = parse_date(first)
                desc = (str(r[1]).strip() if len(r)>1 and r[1] else None)
                acct_name = r[3] if len(r) > 3 else None
                if src == "OfficeExpense":
                    # Office: Date | Description | Income | Account Name | Qty | Price | Amount
                    acct_name = r[3] if len(r) > 3 else None
                    income_cash = to_int_money(r[2]) if len(r) > 2 else 0
                    expense_cash = to_int_money(r[6]) if len(r) > 6 else 0
                    income_kpay = 0
                    expense_kpay = 0
                    qty = r[4] if len(r) > 4 else None
                    unit = to_int_money(r[5]) if len(r) > 5 else None
                else:
                    income_cash  = to_int_money(r[4]) if len(r)>4 else 0
                    income_kpay  = to_int_money(r[5]) if len(r)>5 else 0
                    expense_cash = to_int_money(r[6]) if len(r)>6 else 0
                    expense_kpay = to_int_money(r[7]) if len(r)>7 else 0
                    qty = unit = None
                if not d and not desc: continue
                if not d: d = parse_date(f"01.{tag.split('-')[1]}.{tag.split('-')[0]}")
                if not d: continue
                acct = lookup_account(acct_name)
                rows.append(
                    f"  ({q(d)}, {q(desc)}, "
                    f"(select id from chart_of_accounts where group_name={q(acct)} limit 1), "
                    f"{income_cash or 0}, {income_kpay or 0}, {expense_cash or 0}, {expense_kpay or 0}, "
                    f"{q(qty) if qty is not None and not isinstance(qty,str) else 'NULL'}, "
                    f"{unit or 'NULL'}, "
                    f"{q(src)}, {q(fp.name)})"
                )
        wb.close()
    if not rows: return ["-- no ledger rows"]
    return [
        "insert into ledger_entries (entry_date, description, account_id, income_cash, income_kpay, expense_cash, expense_kpay, qty, unit_price, source, source_file) values\n"
        + ",\n".join(rows) + ";"
    ]

# ─── 5. Events ─────────────────────────────────────────────────────────────
def gen_events() -> list[str]:
    rows_e = []
    rows_i = []
    for name, fp, dt in EVENT_FILES:
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        ws = wb[wb.sheetnames[0]]
        cells = list(ws.iter_rows(values_only=True))
        # Total often in R1C4 or similar
        budget = None
        for r in cells[:3]:
            for v in r or []:
                m = to_int_money(v)
                if m and m > 100000:
                    budget = m; break
            if budget: break
        rows_e.append(
            f"  ({q(name)}, {q(dt)}, {q(budget)}, NULL)"
        )
        # Items: skip rows without an item description or amount
        for r in cells[2:]:
            if not r or not r[0]: continue
            item = str(r[0]).strip()
            if not item or item.lower() in ("qty","amount"): continue
            qty = r[1] if len(r)>1 else None
            unit = to_int_money(r[2]) if len(r)>2 else None
            amt = to_int_money(r[3]) if len(r)>3 else None
            supplier = str(r[4]).strip() if len(r)>4 and r[4] else None
            if amt is None and unit is None: continue
            rows_i.append(
                "  ("
                f"(select id from events where name={q(name)} limit 1),"
                f"{q(item)}, {q(qty) if qty is not None and not isinstance(qty,str) else 'NULL'}, "
                f"{unit or 'NULL'}, {amt or 'NULL'}, {q(supplier)}, TRUE"
                ")"
            )
        wb.close()
    out = []
    if rows_e:
        out.append("insert into events (name, event_date, budget, actual_cost) values\n"
                   + ",\n".join(rows_e) + "\non conflict do nothing;")
    if rows_i:
        out.append("insert into event_budget_items (event_id, item, qty, unit_price, amount, supplier_name, is_estimate) values\n"
                   + ",\n".join(rows_i) + ";")
    return out or ["-- no events"]


# ─── Orchestrator ──────────────────────────────────────────────────────────
def main():
    print("ETL → SQL files in", OUT)
    # 1. Fee schedule
    sql_fee = gen_fee_schedule()
    (OUT / "01_fee_schedule.sql").write_text("BEGIN;\n" + "\n".join(sql_fee) + "\nCOMMIT;\n")
    print("  ✓ 01_fee_schedule.sql")

    # 2-9. Students/sections/etc.
    files = gen_students_attendance()
    am_chunks = files.pop("__ATTENDANCE_CHUNKS__", [])
    for name, lines in files.items():
        (OUT / name).write_text("BEGIN;\n" + "\n".join(lines) + "\nCOMMIT;\n")
        print(f"  ✓ {name}")
    # Attendance: one SQL file per chunk
    for i, stmt in enumerate(am_chunks, 1):
        fp = OUT / f"09_attendance_marks_{i:03d}.sql"
        fp.write_text("BEGIN;\n" + stmt + "\nCOMMIT;\n")
    print(f"  ✓ 09_attendance_marks_*.sql ({len(am_chunks)} chunks)")

    # 10. KPay
    (OUT / "10_kpay_transactions.sql").write_text("BEGIN;\n" + "\n".join(gen_kpay_transactions()) + "\nCOMMIT;\n")
    print("  ✓ 10_kpay_transactions.sql")

    # 11. Ledger
    (OUT / "11_ledger_entries.sql").write_text("BEGIN;\n" + "\n".join(gen_ledger_entries()) + "\nCOMMIT;\n")
    print("  ✓ 11_ledger_entries.sql")

    # 12. Events
    (OUT / "12_events.sql").write_text("BEGIN;\n" + "\n".join(gen_events()) + "\nCOMMIT;\n")
    print("  ✓ 12_events.sql")

    print("\nDone. SQL files ready in:", OUT)

if __name__ == "__main__":
    main()
