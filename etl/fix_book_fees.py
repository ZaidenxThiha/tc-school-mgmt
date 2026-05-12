#!/usr/bin/env python3
"""Scan all daily-income source files for Book Fee and ensure each is recorded as an invoice_line."""
from __future__ import annotations
import os, re, datetime, urllib.parse
from pathlib import Path
import openpyxl
import psycopg2

ROOT = Path(__file__).resolve().parent.parent
PW  = urllib.parse.quote_plus(os.environ.get("TNC_DB_PW", "Thihaaung1@"))
URL = os.environ.get("TNC_DB_URL",
    f"postgresql://postgres.ugjujibpbasskampuums:{PW}@aws-1-ap-southeast-2.pooler.supabase.com:5432/postgres"
)

FILES = [
    ("2026-01", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 Jan ESL Opening Daily Cash Income.xlsx", "Daily Transcation"),
    ("2026-02", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 Feb  ESL  Daily Income Cash.xlsx",       "Daily Transcation"),
    ("2026-03", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/March ESL Daily Income.xlsx",                 "Daily"),
    ("2026-04", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/April ESL Daily Income.xlsx",                 "Daily"),
    ("2026-01", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 Jan  Daily K Pay Transcation.xlsx",  "Daily Income K Pay"),
    ("2026-02", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 Feb  Daily K Pay Transcation.xlsx",  "Daily Income K Pay"),
    ("2026-03", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 March  Daily K Pay Transcation.xlsx", "Daily Income K Pay"),
    ("2026-04", ROOT / "2026 Thazin&Cherry Finance/ESL Daily Income ( Cash , Q Pay)/2026 April  Daily K Pay Transcation.xlsx", "Daily Income K Pay"),
]

def norm(s): return re.sub(r"\s+", " ", str(s)).strip().lower() if s else ""

def main():
    conn = psycopg2.connect(URL, connect_timeout=15)
    cur  = conn.cursor()

    # Build name → student_id lookup
    cur.execute("select id, english_name, myanmar_name from students")
    name_to_id = {}
    for sid, en, mm in cur.fetchall():
        for n in (en, mm):
            k = norm(n)
            if k and k != "-" and k not in name_to_id:
                name_to_id[k] = sid

    # Aggregate source: (student_id, month) → expected book total
    expected = {}
    for tag, fp, sheet in FILES:
        if not fp.exists(): continue
        wb = openpyxl.load_workbook(fp, data_only=True)
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr = None; has_id = False
        for i, r in enumerate(rows[:6]):
            if r and r[0] and isinstance(r[0], str) and r[0].strip().lower() == 'date':
                hdr = i; has_id = any(c and 'id code' in str(c).lower() for c in r); break
        if hdr is None: continue
        mm_col = 2 if has_id else 1
        en_col = 3 if has_id else 2
        book_col = 6 if has_id else 5
        month_first = datetime.date(int(tag.split('-')[0]), int(tag.split('-')[1]), 1)
        for r in rows[hdr+1:]:
            if not r or not r[0]: continue
            if isinstance(r[0], str) and 'total' in r[0].lower(): continue
            mm = r[mm_col] if len(r) > mm_col else None
            en = r[en_col] if len(r) > en_col else None
            try: book = r[book_col] if len(r) > book_col else None
            except (IndexError, TypeError): book = None
            if not isinstance(book, (int, float)) or book <= 0: continue
            sid = name_to_id.get(norm(en)) or name_to_id.get(norm(mm))
            if not sid: continue
            key = (sid, month_first)
            expected[key] = expected.get(key, 0) + int(book)
        wb.close()

    # Aggregate current DB
    cur.execute("""
        select i.student_id, i.billing_month, coalesce(sum(il.amount), 0)::bigint
        from invoices i
        left join invoice_lines il on il.invoice_id = i.id and il.kind = 'book'
        where i.status = 'paid' and i.billing_month between '2026-01-01' and '2026-04-30'
        group by 1, 2
    """)
    current = {(sid, m): int(amt) for sid, m, amt in cur.fetchall()}

    added = 0
    added_amount = 0
    skipped = 0
    for (sid, month), exp_amt in expected.items():
        cur_amt = current.get((sid, month), 0)
        diff = exp_amt - cur_amt
        if diff <= 0:
            continue
        # Find an invoice for this student+month, or create one
        cur.execute("""
            select id from invoices where student_id=%s and billing_month=%s
              and status != 'void' order by id limit 1
        """, (sid, month))
        row = cur.fetchone()
        if row:
            inv_id = row[0]
            # Add the missing book line
            cur.execute("""
                insert into invoice_lines (invoice_id, kind, amount)
                values (%s, 'book', %s)
            """, (inv_id, diff))
            # Bump invoice total
            cur.execute("update invoices set total_amount = total_amount + %s where id = %s", (diff, inv_id))
            added += 1
            added_amount += diff
        else:
            skipped += 1
    conn.commit()
    print(f"Added {added} missing book lines totaling {added_amount:,} MMK")
    print(f"Skipped {skipped} (no matching invoice)")

    # Verify
    cur.execute("""
      select to_char(sum(amount), 'FM999,999,999') from invoice_lines il
      join invoices i on i.id=il.invoice_id where il.kind='book' and i.status='paid'
    """)
    print(f"New DB Book Fee total: {cur.fetchone()[0]} MMK")
    conn.close()

if __name__ == "__main__":
    main()
