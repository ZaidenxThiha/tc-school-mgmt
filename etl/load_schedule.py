#!/usr/bin/env python3
"""Load class schedule (room × time-slot × Sat/Sun, with MT and CT) for Feb/Mar/Apr 2026."""
from __future__ import annotations
import os, re, datetime, urllib.parse
from pathlib import Path
import openpyxl
import psycopg2

ROOT = Path(__file__).resolve().parent.parent
HR   = ROOT / "2026 Thazin&Cherry HR"
# Prefer newer copy at the project root if it exists, else the HR folder copy
_ROOT_FILE = ROOT / "_Official 2026 Thazin&Cherry Class schedule Template.xlsx"
SCHEDULE_FILE = _ROOT_FILE if _ROOT_FILE.exists() else HR / " Official 2026 Thazin&Cherry Class schedule Template.xlsx"

PW  = urllib.parse.quote_plus(os.environ.get("TNC_DB_PW", "Thihaaung1@"))
URL = os.environ.get("TNC_DB_URL",
    f"postgresql://postgres.ugjujibpbasskampuums:{PW}@"
    f"aws-1-ap-southeast-2.pooler.supabase.com:5432/postgres"
)

# Column structure (0-indexed): cols 2-9 are 4 slots × Sat/Sun
SLOTS = [
    ('7:45-9:45', 2, 3),
    ('10-12',     4, 5),
    ('1-3',       6, 7),
    ('3:15-5:15', 8, 9),
]

LEVEL_FROM_LABEL = {
    'EARLYCHILDHOOD': 'EARLY_CHILDHOOD',
    'EARLY CHILDHOOD': 'EARLY_CHILDHOOD',
    'NURSERY': 'NURSERY',
    'PRESTARTER': 'PRE_STARTER', 'PRE-STARTER': 'PRE_STARTER', 'PRE STARTER': 'PRE_STARTER',
    'PRESTARTERS': 'PRE_STARTER',
    'STARTER': 'STARTER', 'STARTERS': 'STARTER',
    'MOVER': 'MOVERS', 'MOVERS': 'MOVERS',
    'FLYER': 'FLYERS', 'FLYERS': 'FLYERS',
    'KEY': 'KEY', 'PET': 'PET', 'FCE': 'FCE', 'CAE': 'CAE',
}

def parse_class_label(label: str):
    """'Nursery (10-12)' or 'Pre-starters Online (10-12)' → (level_code, time_slot, is_online)."""
    if not label: return (None, None, False)
    s = label.strip()
    is_online = bool(re.search(r"\bonline\b", s, re.I))
    s = re.sub(r"\bonline\b", "", s, flags=re.I).strip()
    m = re.match(r"^([A-Za-z\- ]+)\s*\(?\s*([\d.:\-\s]*)\s*\)?\s*$", s)
    if not m: return (None, None, is_online)
    level_word = m.group(1).strip().upper().replace('-', '')
    slot = m.group(2).strip().replace(' ', '') or None
    code = LEVEL_FROM_LABEL.get(level_word.replace(' ',''))
    return (code, slot, is_online)

def load_month(conn, sheet_name: str, month: datetime.date):
    cur = conn.cursor()
    wb = openpyxl.load_workbook(SCHEDULE_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  ! {sheet_name} not in workbook"); return
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Wipe existing assignments for this month
    cur.execute("delete from schedule_assignments where month = %s", (month,))

    # Build employee + room + section lookups
    cur.execute("select id, lower(replace(short_name, ' ', '')) from employees")
    emp_lookup = {k: v for v, k in cur.fetchall()}
    def find_emp(name: str | None):
        if not name: return None
        key = re.sub(r'\s+', '', name).lower()
        if key in emp_lookup: return emp_lookup[key]
        # Try fuzzy: drop "Tr." prefix
        key2 = re.sub(r'^tr\.?', '', key)
        return emp_lookup.get('tr' + key2)

    cur.execute("select id, lower(name) from rooms")
    room_lookup = {k: v for v, k in cur.fetchall()}
    def find_room(name: str | None):
        if not name: return None
        return room_lookup.get(name.strip().lower())

    cur.execute("""
        select s.id, l.code, s.time_slot, s.is_online
        from sections s join levels l on l.id = s.level_id
    """)
    sec_lookup = {(c, t, o): sid for sid, c, t, o in cur.fetchall()}

    # Walk room blocks (each block = 4 rows: Class Name, Subject, MT, CT)
    inserted = 0
    i = 2  # data starts row 3 (index 2)
    while i < len(rows) - 3:
        r = rows[i]
        if r and r[0] and isinstance(r[0], str) and r[0].strip().lower().startswith('room'):
            room_label = r[0].strip()
            room_id = find_room(room_label)
            class_row   = rows[i]      # row N: 'Class Name' + class labels
            subject_row = rows[i+1]    # row N+1: 'Subject' + subjects
            mt_row      = rows[i+2]    # row N+2: 'MT'
            ct_row      = rows[i+3]    # row N+3: 'CT'

            for slot_label, sat_col, sun_col in SLOTS:
                for day, col in (('Sat', sat_col), ('Sun', sun_col)):
                    if col >= len(class_row): continue
                    class_label = class_row[col]
                    subject = subject_row[col] if subject_row and col < len(subject_row) else None
                    mt_name = mt_row[col]      if mt_row      and col < len(mt_row)      else None
                    ct_name = ct_row[col]      if ct_row      and col < len(ct_row)      else None

                    # Skip cells with no class assigned
                    if not class_label and not mt_name and not ct_name: continue

                    level_code, parsed_slot, is_online = parse_class_label(str(class_label) if class_label else '')
                    section_id = None
                    if level_code:
                        # Use parsed slot if present, else slot_label
                        section_id = sec_lookup.get((level_code, parsed_slot or slot_label, is_online))
                        if section_id is None:
                            section_id = sec_lookup.get((level_code, slot_label, is_online))

                    cur.execute("""
                        insert into schedule_assignments (
                          month, day_of_week, time_slot, room_id, section_id,
                          class_label, subject, mt_employee_id, ct_employee_id
                        ) values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        month, day, slot_label, room_id, section_id,
                        str(class_label).strip() if class_label else None,
                        str(subject).strip() if subject else None,
                        find_emp(str(mt_name) if mt_name else None),
                        find_emp(str(ct_name) if ct_name else None),
                    ))
                    inserted += 1
            i += 4
        else:
            i += 1
    wb.close()
    conn.commit()
    print(f"  {sheet_name}: {inserted} assignments")

def main():
    conn = psycopg2.connect(URL, connect_timeout=15)
    print(f"Loading schedule from {SCHEDULE_FILE.name} …")
    for sheet, month in (
        ('PlannerTemplateTimetable', datetime.date(2026, 1, 1)),
        ('February Timetable',       datetime.date(2026, 2, 1)),
        ('March Timetable',          datetime.date(2026, 3, 1)),
        ('April Timetable',          datetime.date(2026, 4, 1)),
        ('May Timetable',            datetime.date(2026, 5, 1)),
        ('June Timetable',           datetime.date(2026, 6, 1)),   # safe to skip if missing
    ):
        load_month(conn, sheet, month)
    conn.close()
    print("Done.")

if __name__ == "__main__":
    main()
