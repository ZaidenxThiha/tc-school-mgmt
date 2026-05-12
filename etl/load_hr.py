#!/usr/bin/env python3
"""
HR ETL — loads teacher profiles, admin staff, rooms, payslips, absences.
"""
from __future__ import annotations
import os, re, sys, datetime, urllib.parse
from pathlib import Path
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

ROOT = Path(__file__).resolve().parent.parent
HR   = ROOT / "2026 Thazin&Cherry HR"

PW  = urllib.parse.quote_plus(os.environ.get("TNC_DB_PW", "Thihaaung1@"))
URL = os.environ.get("TNC_DB_URL",
    f"postgresql://postgres.ugjujibpbasskampuums:{PW}@"
    f"aws-1-ap-southeast-2.pooler.supabase.com:5432/postgres"
)

def short_from_full(full: str) -> str:
    """'Aye Cherry Tun@Tr Cherry' → 'Tr Cherry'"""
    if not full: return ''
    if '@' in full:
        return full.split('@', 1)[1].strip()
    parts = full.strip().split()
    if not parts: return full.strip()
    if parts[0].lower() in ('tr', 'tr.'):
        return ' '.join(parts[:2])
    return f"Tr {parts[0]}"

def to_int(v):
    if v is None or v == '' or v == '-': return None
    if isinstance(v, (int, float)):
        try: return int(v)
        except: return None
    s = str(v).strip().replace(',', '')
    try: return int(float(s))
    except: return None

def to_num(v):
    if v is None or v == '' or v == '-': return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip().replace(',', ''))
    except: return None

def parse_date(v):
    if v is None or v == '': return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d.%m.%y', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

# ──────────────────────────────────────────────────────────────────────────
def load_profiles(conn):
    cur = conn.cursor()
    fp = HR / "2026 T&C Cambridge Program Teaching Staff Info (Responses).xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["Form Responses 1"]
    rows = list(ws.iter_rows(values_only=True))
    inserted, updated = 0, 0
    for r in rows[1:]:
        if not r or not r[1]: continue
        full = str(r[1]).strip()
        short = short_from_full(full)
        dob = parse_date(r[2]) if len(r) > 2 else None
        address = str(r[3]).strip() if len(r) > 3 and r[3] else None
        nrc = str(r[4]).strip() if len(r) > 4 and r[4] else None
        phone1 = str(r[5]).strip() if len(r) > 5 and r[5] else None
        phone2 = str(r[6]).strip() if len(r) > 6 and r[6] and str(r[6]).lower() != 'none' else None
        email = str(r[7]).strip() if len(r) > 7 and r[7] else None
        edu = str(r[8]).strip() if len(r) > 8 and r[8] else None
        deg_uni = str(r[9]).strip() if len(r) > 9 and r[9] else None
        emergency = str(r[10]).strip() if len(r) > 10 and r[10] else None
        position = str(r[11]).strip() if len(r) > 11 and r[11] else 'Main Teacher'
        full_time = str(r[12]).strip() if len(r) > 12 and r[12] else 'Yes'
        slots = str(r[13]).strip() if len(r) > 13 and r[13] else None

        # Try match existing employee by short_name (case/space-insensitive)
        cur.execute("""
            select id from employees
             where lower(replace(short_name, '.', '')) = lower(replace(%s, '.', ''))
                or lower(replace(short_name, ' ', '')) = lower(replace(%s, ' ', ''))
            limit 1
        """, (short, short))
        row = cur.fetchone()

        avail = ('Full-time' if full_time.lower() == 'yes' else 'Part-time') + \
                (f" — {slots}" if slots and slots.lower() != 'none' else '')

        if row:
            cur.execute("""
                update employees set
                  full_name = %s, date_of_birth = %s, national_id = %s,
                  address = coalesce(%s, address),
                  phone = coalesce(%s, phone),
                  email = coalesce(%s, email),
                  emergency_contact = %s,
                  position_title = %s,
                  education_level = %s, degree = %s,
                  available_slots = %s
                where id = %s
            """, (full, dob, nrc, address, phone1, email, emergency, position, edu, deg_uni, avail, row[0]))
            updated += 1
        else:
            cur.execute("""
                insert into employees (
                  short_name, full_name, category, date_of_birth, national_id,
                  address, phone, email, emergency_contact, position_title,
                  education_level, degree, available_slots, is_active
                ) values (%s,%s,'esl_teacher',%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,true)
            """, (short, full, dob, nrc, address, phone1, email, emergency, position, edu, deg_uni, avail))
            inserted += 1
    wb.close()
    conn.commit()
    print(f"  profiles: inserted {inserted}, updated {updated}")

# ──────────────────────────────────────────────────────────────────────────
def load_admin_staff(conn):
    cur = conn.cursor()
    fp = HR / "2026 T&C Salary Payment.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    ws = wb["Admin"]
    rows = list(ws.iter_rows(values_only=True))
    # R1: [None, 'February', 'March', 'April']
    inserted = 0
    for r in rows[1:]:
        if not r or not r[0]: continue
        name = str(r[0]).strip()
        if name.lower().startswith('total'): continue
        # Use feb amount as monthly salary
        feb = to_int(r[1]) if len(r) > 1 else None
        cur.execute("""
            insert into employees (short_name, full_name, category, monthly_salary, is_active)
            values (%s, %s, 'admin_staff', %s, true)
            on conflict do nothing
            returning id
        """, (name, name, feb))
        if cur.fetchone():
            inserted += 1
    wb.close()
    conn.commit()
    print(f"  admin_staff: inserted {inserted}")

# ──────────────────────────────────────────────────────────────────────────
def load_rooms(conn):
    """Extract distinct rooms from the schedule template."""
    cur = conn.cursor()
    fp = HR / " Official 2026 Thazin&Cherry Class schedule Template.xlsx"
    if not fp.exists():
        # fallback to root copy
        fp = ROOT / "_Official 2026 Thazin&Cherry Class schedule Template.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    rooms_seen: set[tuple[str, str | None]] = set()
    for sheet in ("PlannerTemplateTimetable",):
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows:
            if not r: continue
            v = r[0]
            if isinstance(v, str) and v.strip().lower().startswith('room'):
                room = v.strip()
                # next row may have a "display name" like 'Pearl' in r[0]
                rooms_seen.add((room, None))
            elif isinstance(v, str) and v.strip() in ('Pearl', 'Diamond', 'Ruby', 'Topaz', 'Jade', 'Amber', 'Coral', 'Crystal'):
                # Single display-name word, attach to last seen
                pass
    # Fallback: scan first column for labels Room X / known names
    if not rooms_seen:
        for sheet in ("PlannerTemplateTimetable",):
            if sheet not in wb.sheetnames: continue
            ws = wb[sheet]
            for r in ws.iter_rows(min_col=1, max_col=2, values_only=True):
                if r and r[0] and isinstance(r[0], str):
                    v = r[0].strip()
                    if v.lower().startswith('room ') or v in ('Pearl','Diamond','Ruby','Topaz','Jade','Amber','Coral','Crystal'):
                        rooms_seen.add((v, None))

    inserted = 0
    for room, dn in sorted(rooms_seen):
        cur.execute("insert into rooms (name, display_name) values (%s, %s) on conflict do nothing returning id",
                    (room, dn))
        if cur.fetchone():
            inserted += 1
    wb.close()
    conn.commit()
    print(f"  rooms: inserted {inserted} (total {len(rooms_seen)} seen)")

# ──────────────────────────────────────────────────────────────────────────
def load_payslips(conn):
    """Load monthly payslips from 2026 Teacher_s salary payment list.xlsx."""
    cur = conn.cursor()
    fp = HR / "2026 Teacher_s salary payment list.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)

    # Month to sheet mapping
    months = [
        (datetime.date(2026, 2, 1), "FebruarySalary"),
        (datetime.date(2026, 3, 1), "MarchSalary"),
        (datetime.date(2026, 4, 1), "AprilSalary"),
    ]
    rate_cache: dict[str, tuple[int|None,int|None]] = {}

    for pay_month, sn in months:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        # header at row 4 (idx 3): No, Teachers, MT Hourly, CT Hourly, Sat MT, Sun MT, Sat CT, Sun CT, MT total, CT total, MT abs, CT abs, Total Pay
        for r in rows[4:]:
            if not r or r[0] is None or not isinstance(r[0], (int, float)): continue
            tname = str(r[1]).strip() if r[1] else None
            if not tname or tname.lower() == 'teachers': continue
            mt_fee = to_int(r[2])
            ct_fee = to_int(r[3])
            mt_total = to_num(r[8]) or 0
            ct_total = to_num(r[9]) or 0
            mt_abs = to_num(r[10]) or 0
            ct_abs = to_num(r[11]) or 0
            total_pay = to_int(r[12]) or 0

            # Lookup employee by short_name (fuzzy)
            cur.execute("""
                select id, mt_hourly_fee, ct_hourly_fee from employees
                 where lower(replace(short_name, ' ', '')) = lower(replace(%s, ' ', ''))
                    or lower(short_name) ilike %s
                limit 1
            """, (tname, f"%{tname.replace('Tr ','')}%"))
            row = cur.fetchone()
            if not row:
                # Insert new teacher
                cur.execute("""
                    insert into employees (short_name, full_name, category, mt_hourly_fee, ct_hourly_fee, is_active)
                    values (%s, %s, 'esl_teacher', %s, %s, true)
                    returning id
                """, (tname, tname, mt_fee, ct_fee))
                emp_id = cur.fetchone()[0]
                rate_cache[tname] = (mt_fee, ct_fee)
            else:
                emp_id, cur_mt, cur_ct = row
                # Update hourly rates if missing
                if (cur_mt is None and mt_fee) or (cur_ct is None and ct_fee):
                    cur.execute("""
                        update employees set
                          mt_hourly_fee = coalesce(mt_hourly_fee, %s),
                          ct_hourly_fee = coalesce(ct_hourly_fee, %s)
                        where id = %s
                    """, (mt_fee, ct_fee, emp_id))

            # ESL pay = (mt hours - mt absence) × mt_fee + (ct hours - ct absence) × ct_fee
            mt_pay = max(0, (mt_total - mt_abs)) * (mt_fee or 0)
            ct_pay = max(0, (ct_total - ct_abs)) * (ct_fee or 0)
            esl_pay = int(mt_pay + ct_pay)

            cur.execute("""
                insert into employee_payslips (
                  employee_id, pay_month,
                  mt_hours, ct_hours, mt_absence_hrs, ct_absence_hrs,
                  mt_hourly_fee, ct_hourly_fee,
                  esl_pay
                ) values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                on conflict (employee_id, pay_month) do update set
                  mt_hours = excluded.mt_hours, ct_hours = excluded.ct_hours,
                  mt_absence_hrs = excluded.mt_absence_hrs, ct_absence_hrs = excluded.ct_absence_hrs,
                  mt_hourly_fee = excluded.mt_hourly_fee, ct_hourly_fee = excluded.ct_hourly_fee,
                  esl_pay = excluded.esl_pay
            """, (emp_id, pay_month, mt_total, ct_total, mt_abs, ct_abs, mt_fee or 0, ct_fee or 0, esl_pay))

    wb.close()

    # Component pay (management/guide/summer) from 2026 T&C Salary Payment.xlsx > Teachers sheet
    fp2 = HR / "2026 T&C Salary Payment.xlsx"
    wb2 = openpyxl.load_workbook(fp2, data_only=True)
    if "Teachers" in wb2.sheetnames:
        ws = wb2["Teachers"]
        rows = list(ws.iter_rows(values_only=True))
        # Header layout: cols 4-8 = February (ESL, Mgmt, Guide, Others, Total),
        # 9-13 = March, 14-18 = April
        month_offsets = [
            (datetime.date(2026, 2, 1), 4),
            (datetime.date(2026, 3, 1), 9),
            (datetime.date(2026, 4, 1), 14),
        ]
        for r in rows[2:]:
            if not r or not r[1]: continue
            tname = str(r[1]).strip()
            if tname.lower() == 'full name': continue
            cur.execute("""
                select id from employees
                 where lower(replace(short_name, ' ', '')) = lower(replace(%s, ' ', ''))
                    or lower(full_name) = lower(%s)
                limit 1
            """, (tname, tname))
            emp = cur.fetchone()
            if not emp: continue
            emp_id = emp[0]
            for pay_month, off in month_offsets:
                if off + 4 >= len(r): continue
                # esl already loaded from PlannerSheet; just attach mgmt/guide/others
                mgmt   = to_int(r[off + 1]) or 0
                guide  = to_int(r[off + 2]) or 0
                others = to_int(r[off + 3]) or 0
                if mgmt + guide + others == 0: continue
                cur.execute("""
                    insert into employee_payslips (employee_id, pay_month,
                      management_pay, guide_pay, other_pay)
                    values (%s, %s, %s, %s, %s)
                    on conflict (employee_id, pay_month) do update set
                      management_pay = excluded.management_pay,
                      guide_pay = excluded.guide_pay,
                      other_pay = excluded.other_pay
                """, (emp_id, pay_month, mgmt, guide, others))

    # Summer pay
    if "Summers" in wb2.sheetnames:
        ws = wb2["Summers"]
        rows = list(ws.iter_rows(values_only=True))
        # Header: [None, 'Year-6 Guide', 'March', 'April']
        for r in rows[1:]:
            if not r or not r[1]: continue
            tname = str(r[1]).strip()
            mar = to_int(r[2]) if len(r) > 2 else None
            apr = to_int(r[3]) if len(r) > 3 else None
            cur.execute("""
                select id from employees
                 where lower(replace(short_name, ' ', '')) = lower(replace(%s, ' ', ''))
                    or lower(full_name) ilike %s
                limit 1
            """, (tname, f"%{tname}%"))
            emp = cur.fetchone()
            if not emp: continue
            emp_id = emp[0]
            for pay_month, amt in ((datetime.date(2026,3,1), mar), (datetime.date(2026,4,1), apr)):
                if not amt: continue
                cur.execute("""
                    insert into employee_payslips (employee_id, pay_month, summer_pay)
                    values (%s, %s, %s)
                    on conflict (employee_id, pay_month) do update set
                      summer_pay = employee_payslips.summer_pay + excluded.summer_pay
                """, (emp_id, pay_month, amt))

    wb2.close()
    conn.commit()
    print(f"  payslips loaded for Feb/Mar/Apr 2026")

# ──────────────────────────────────────────────────────────────────────────
def load_absences(conn):
    cur = conn.cursor()
    fp = HR / "2026 Teacher_s salary payment list.xlsx"
    wb = openpyxl.load_workbook(fp, data_only=True)
    if "Absence-Log" not in wb.sheetnames: return
    ws = wb["Absence-Log"]
    rows = list(ws.iter_rows(values_only=True))
    inserted = 0
    for r in rows[1:]:
        if not r or len(r) < 7: continue
        d = parse_date(r[1])
        if not d: continue
        tname  = str(r[4]).strip() if r[4] else None
        if not tname: continue
        role   = (str(r[5]).strip() if r[5] else 'MT')
        # Convert role text to MT/CT
        if 'main' in role.lower(): role = 'MT'
        elif 'class' in role.lower() or role.upper() == 'CT': role = 'CT'
        else: role = 'MT'
        hours_text = str(r[6]).strip() if r[6] else ''
        # parse hours from text like "10:00 - 12:00 (2 hrs)"
        m = re.search(r"\((\d+(?:\.\d+)?)\s*hrs?\)", hours_text)
        hours = float(m.group(1)) if m else 1.0
        reason = str(r[8]).strip() if len(r) > 8 and r[8] else None
        # Find employee
        cur.execute("""
            select id from employees
             where lower(replace(short_name, ' ', '')) = lower(replace(%s, ' ', ''))
                or lower(full_name) ilike %s
            limit 1
        """, (tname, f"%{tname.replace('Tr ','')}%"))
        emp = cur.fetchone()
        if not emp: continue
        cur.execute("""
            insert into absences (employee_id, absent_date, hours, role, reason)
            values (%s, %s, %s, %s, %s)
        """, (emp[0], d, hours, role, reason))
        inserted += 1
    wb.close()
    conn.commit()
    print(f"  absences: inserted {inserted}")

# ──────────────────────────────────────────────────────────────────────────
def main():
    conn = psycopg2.connect(URL, connect_timeout=15)
    print("Loading HR data → employees / rooms / payslips / absences …")
    load_profiles(conn)
    load_admin_staff(conn)
    load_rooms(conn)
    load_payslips(conn)
    load_absences(conn)
    conn.close()
    print("Done.")

if __name__ == "__main__":
    main()
