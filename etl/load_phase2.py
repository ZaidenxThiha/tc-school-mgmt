#!/usr/bin/env python3
"""
Phase 2 ETL — promotes daily-income files into invoices + payments + ledger,
and loads ESL/Summer opening cost files into products + purchase_orders +
po_items + inventory_movements.

Connects directly via psycopg2 using TNC_DB_URL or default pooler URL.
"""
from __future__ import annotations
import os, re, sys, hashlib, datetime, urllib.parse
from pathlib import Path
import openpyxl
import psycopg2

ROOT = Path(__file__).resolve().parent.parent
FIN  = ROOT / "2026 Thazin&Cherry Finance"

PW = urllib.parse.quote_plus(os.environ.get("TNC_DB_PW", "Thihaaung1@"))
URL = os.environ.get("TNC_DB_URL",
    f"postgresql://postgres.ugjujibpbasskampuums:{PW}@"
    f"aws-1-ap-southeast-2.pooler.supabase.com:5432/postgres"
)

# ─── Files ─────────────────────────────────────────────────────────────────
CASH_FILES = [
    ("2026-01", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Jan ESL Opening Daily Cash Income.xlsx", "Daily Transcation"),
    ("2026-02", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Feb  ESL  Daily Income Cash.xlsx",       "Daily Transcation"),
    ("2026-03", FIN / "ESL Daily Income ( Cash , Q Pay)/March ESL Daily Income.xlsx",                 "Daily"),
    ("2026-04", FIN / "ESL Daily Income ( Cash , Q Pay)/April ESL Daily Income.xlsx",                 "Daily"),
]
KPAY_FILES = [
    ("2026-01", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Jan  Daily K Pay Transcation.xlsx",  "Daily Income K Pay"),
    ("2026-02", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 Feb  Daily K Pay Transcation.xlsx",  "Daily Income K Pay"),
    ("2026-03", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 March  Daily K Pay Transcation.xlsx", "Daily Income K Pay"),
    ("2026-04", FIN / "ESL Daily Income ( Cash , Q Pay)/2026 April  Daily K Pay Transcation.xlsx", "Daily Income K Pay"),
]
ESL_OPENING = FIN / "T&C All Program Opening Cost/2026 ESL Opening Cost.xlsx"
SUMMER_OPENING = FIN / "T&C All Program Opening Cost/2026 Summer Opening Cost.xlsx"

LEVEL_FROM_LABEL = {
    "EARLYCHILDHOOD": "EARLY_CHILDHOOD",
    "EARLY CHILDHOOD": "EARLY_CHILDHOOD",
    "NURSERY": "NURSERY",
    "PRESTARTER": "PRE_STARTER", "PRE-STARTER": "PRE_STARTER", "PRE STARTER": "PRE_STARTER",
    "STARTER": "STARTER", "STARTERS": "STARTER",
    "MOVER": "MOVERS", "MOVERS": "MOVERS",
    "FLYER": "FLYERS", "FLYERS": "FLYERS",
    "KEY": "KEY",
    "PET": "PET",
    "FCE": "FCE",
    "CAE": "CAE",
}

def parse_class(label: str) -> tuple[str | None, str | None, bool]:
    """Parse 'Flyers (3:15-5:15)' or 'PET Online (10-12)' → (level_code, time_slot, is_online)."""
    if not label: return (None, None, False)
    s = label.strip()
    is_online = bool(re.search(r"\bonline\b", s, re.I))
    s = re.sub(r"\bonline\b", "", s, flags=re.I).strip()
    # Take everything before "(" as level, inside parens as slot
    m = re.match(r"^([A-Za-z\- ]+)\s*\(?\s*([\d.:\-\s]*)\s*\)?\s*$", s)
    if not m: return (None, None, is_online)
    level_word = m.group(1).strip().upper()
    slot = m.group(2).strip().replace(" ", "") or None
    code = LEVEL_FROM_LABEL.get(level_word) or LEVEL_FROM_LABEL.get(level_word.replace("-",""))
    if slot:
        # Normalize "745-945" / "315-515" → with colons
        m2 = re.fullmatch(r"(\d{1,2})(\d{2})-(\d{1,2})(\d{2})", slot)
        if m2: slot = f"{int(m2.group(1))}:{m2.group(2)}-{int(m2.group(3))}:{m2.group(4)}"
    return (code, slot, is_online)

def parse_date(v):
    if v is None or v == "": return None
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None

def to_int(v):
    if v is None or v == "" or v == "-": return 0
    if isinstance(v, (int, float)):
        try: return int(v)
        except: return 0
    s = str(v).strip().replace(",", "").replace("MMK","").strip()
    try: return int(float(s))
    except: return 0

def norm_name(s):
    if not s: return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()

# ─── Loaders ───────────────────────────────────────────────────────────────
def build_lookups(cur):
    """Return dicts: name→student_id, (level,slot,online)→section_id."""
    cur.execute("""
        select s.id, s.english_name, s.myanmar_name
        from students s
        where s.english_name is not null or s.myanmar_name is not null
    """)
    name_to_id: dict[str, int] = {}
    for sid, en, mm in cur.fetchall():
        for n in (en, mm):
            k = norm_name(n)
            if k and k != "-" and k not in name_to_id:
                name_to_id[k] = sid

    cur.execute("""
        select se.id, l.code, se.time_slot, se.is_online
        from sections se join levels l on l.id = se.level_id
    """)
    section_to_id: dict[tuple[str,str,bool], int] = {
        (code, slot, is_online): sid
        for sid, code, slot, is_online in cur.fetchall()
    }

    cur.execute("select id from chart_of_accounts where group_name='ESL Class Fee'")
    income_acct = cur.fetchone()[0]
    return name_to_id, section_to_id, income_acct

def load_daily_income(conn):
    cur = conn.cursor()
    name_to_id, section_to_id, income_acct = build_lookups(cur)
    print(f"  lookups: {len(name_to_id)} names, {len(section_to_id)} sections")

    # Wipe existing invoices/payments/lines from prior runs (idempotent)
    cur.execute("delete from invoice_lines; delete from payments; delete from invoices;")
    # Also wipe prior income_cash/income_kpay rows under ESL Class Fee with source='Auto'
    cur.execute("delete from ledger_entries where source='Auto' and account_id=%s", (income_acct,))

    invoices = 0
    payments = 0
    lines = 0
    skipped_no_student = 0
    skipped_no_section = 0
    ledger_added = 0

    for channel, files in (("cash", CASH_FILES), ("kpay", KPAY_FILES)):
        for tag, fp, sheet in files:
            if not fp.exists():
                print(f"  ! missing {fp.name}"); continue
            wb = openpyxl.load_workbook(fp, data_only=True)
            if sheet not in wb.sheetnames:
                print(f"  ! sheet {sheet!r} missing in {fp.name}"); wb.close(); continue
            ws = wb[sheet]
            month_first = datetime.date(int(tag.split('-')[0]), int(tag.split('-')[1]), 1)
            rows_count = 0

            # Detect header row → identify Student ID Code column (some files have it)
            rows = list(ws.iter_rows(values_only=True))
            header_row_idx = None
            has_id_code = False
            for i, r in enumerate(rows[:6]):
                if r and r[0] and str(r[0]).strip().lower() == 'date':
                    header_row_idx = i
                    has_id_code = any(
                        c and 'id code' in str(c).lower() for c in r
                    )
                    break
            if header_row_idx is None: continue

            for r in rows[header_row_idx + 1:]:
                if not r: continue
                d = parse_date(r[0])
                if not d: continue
                # Skip totals/aggregate rows
                if r[0] and isinstance(r[0], str) and 'total' in r[0].lower(): continue

                if has_id_code:
                    mm = r[2] if len(r) > 2 else None
                    en = r[3] if len(r) > 3 else None
                    cls = r[4] if len(r) > 4 else None
                    fees_off = 5
                else:
                    mm = r[1] if len(r) > 1 else None
                    en = r[2] if len(r) > 2 else None
                    cls = r[3] if len(r) > 3 else None
                    fees_off = 4

                esl_fee   = to_int(r[fees_off + 0]) if len(r) > fees_off + 0 else 0
                book_fee  = to_int(r[fees_off + 1]) if len(r) > fees_off + 1 else 0
                id_card   = to_int(r[fees_off + 2]) if len(r) > fees_off + 2 else 0
                tshirt    = to_int(r[fees_off + 3]) if len(r) > fees_off + 3 else 0
                fine_disc = to_int(r[fees_off + 4]) if len(r) > fees_off + 4 else 0
                guide_fee = to_int(r[fees_off + 5]) if len(r) > fees_off + 5 else 0
                amount    = to_int(r[fees_off + 6]) if len(r) > fees_off + 6 else 0

                if amount <= 0: continue

                # Find section
                level_code, slot, online = parse_class(str(cls) if cls else "")
                section_id = section_to_id.get((level_code, slot, online)) if level_code and slot else None
                if section_id is None:
                    skipped_no_section += 1

                # Find student
                student_id = None
                for n in (en, mm):
                    sid = name_to_id.get(norm_name(n))
                    if sid:
                        student_id = sid; break
                if student_id is None:
                    skipped_no_student += 1
                    # Still log income to ledger so revenue totals are accurate
                    cur.execute("""
                        insert into ledger_entries (entry_date, description, account_id,
                            income_cash, income_kpay, source, source_file)
                        values (%s, %s, %s, %s, %s, 'Auto', %s)
                    """, (d, f"Daily {channel} (unmatched): {en or mm} — {cls or ''}",
                          income_acct,
                          amount if channel=='cash' else 0,
                          amount if channel=='kpay' else 0,
                          fp.name))
                    ledger_added += 1
                    continue

                # Insert invoice
                cur.execute("""
                    insert into invoices (student_id, section_id, billing_month, total_amount, status)
                    values (%s, %s, %s, %s, 'paid')
                    returning id
                """, (student_id, section_id, month_first, amount))
                inv_id = cur.fetchone()[0]
                invoices += 1

                # Lines
                line_specs = [
                    ('class_fee', esl_fee), ('book', book_fee), ('id', id_card),
                    ('tshirt', tshirt), ('discount', -abs(fine_disc) if fine_disc else 0),
                    ('guide', guide_fee),
                ]
                for kind, amt in line_specs:
                    if amt:
                        cur.execute("""
                            insert into invoice_lines (invoice_id, kind, amount)
                            values (%s, %s, %s)
                        """, (inv_id, kind, amt))
                        lines += 1

                # Payment
                cur.execute("""
                    insert into payments (invoice_id, student_id, paid_at, amount, channel)
                    values (%s, %s, %s, %s, %s)
                """, (inv_id, student_id, d, amount, channel))
                payments += 1

                # Auto ledger income entry
                cur.execute("""
                    insert into ledger_entries (entry_date, description, account_id,
                        income_cash, income_kpay, source, source_file)
                    values (%s, %s, %s, %s, %s, 'Auto', %s)
                """, (d, f"Daily {channel}: {en or mm} — {cls or ''}",
                      income_acct,
                      amount if channel=='cash' else 0,
                      amount if channel=='kpay' else 0,
                      fp.name))
                ledger_added += 1
                rows_count += 1
            wb.close()
            print(f"    ✓ {fp.name}: {rows_count} rows")

    conn.commit()
    print(f"  invoices={invoices}  payments={payments}  lines={lines}  ledger_added={ledger_added}")
    print(f"  skipped: no_student={skipped_no_student}  no_section={skipped_no_section}")

# ─── Opening costs ─────────────────────────────────────────────────────────
def get_or_create_supplier(cur, name: str) -> int:
    cur.execute("select id from suppliers where name=%s", (name,))
    r = cur.fetchone()
    if r: return r[0]
    cur.execute("insert into suppliers (name) values (%s) returning id", (name,))
    return cur.fetchone()[0]

def get_or_create_product(cur, kind: str, name: str, level_code: str | None,
                          size: str | None, cost: int | None, retail: int | None) -> int:
    level_id = None
    if level_code:
        cur.execute("select id from levels where code=%s", (level_code,))
        r = cur.fetchone()
        if r: level_id = r[0]
    cur.execute("""
        select id from products where kind=%s and name=%s
          and coalesce(level_id, -1) = coalesce(%s, -1)
          and coalesce(size, '') = coalesce(%s, '')
        limit 1
    """, (kind, name, level_id, size))
    r = cur.fetchone()
    if r: return r[0]
    cur.execute("""
        insert into products (kind, name, level_id, size, cost_price, retail_price)
        values (%s, %s, %s, %s, %s, %s) returning id
    """, (kind, name, level_id, size, cost, retail))
    return cur.fetchone()[0]

def load_opening_costs(conn):
    cur = conn.cursor()
    # Wipe prior opening cost data so re-runs are clean
    cur.execute("delete from inventory_movements where reason='purchase';")
    cur.execute("delete from po_items;")
    cur.execute("delete from purchase_orders;")
    cur.execute("delete from products;")
    cur.execute("delete from suppliers;")

    if not ESL_OPENING.exists():
        print("  ! ESL opening cost file missing"); return

    wb = openpyxl.load_workbook(ESL_OPENING, data_only=True)
    products_added = 0
    pos_added = 0
    items_added = 0
    movements_added = 0

    # Textbook orders: 'Textbook Cost' (1st), '2nd order', '3rd order'
    LEVEL_BY_NAME = {
        "early childhood":"EARLY_CHILDHOOD","nursery":"NURSERY",
        "pre-starter":"PRE_STARTER","pre starter":"PRE_STARTER","prestarter":"PRE_STARTER",
        "starter":"STARTER","starters":"STARTER",
        "mover":"MOVERS","movers":"MOVERS",
        "flyer":"FLYERS","flyers":"FLYERS",
        "key":"KEY","pet":"PET","fce":"FCE","cae":"CAE",
    }

    for sheet_name in ("Textbook Cost", "2nd order", "3rd order"):
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Header at row 2: No, Book Name, Level Class, Qty, Price, Book Shop Price, Order Book Shop Name
        order_label = sheet_name
        po_supplier_groups: dict[str, list[tuple]] = {}  # supplier → [(product_id, qty, unit_cost, amount)]
        for r in rows[2:]:
            if not r or not r[0]: continue
            if not isinstance(r[0], (int, float)): continue
            book_name = str(r[1]).strip() if len(r) > 1 and r[1] else None
            level_word = str(r[2]).strip().lower() if len(r) > 2 and r[2] else ""
            qty = r[3] if len(r) > 3 else None
            unit_price = to_int(r[4]) if len(r) > 4 else 0
            shop_price = to_int(r[5]) if len(r) > 5 else 0
            supplier_name = str(r[6]).strip() if len(r) > 6 and r[6] else "Unknown"
            if not book_name or not qty: continue
            level_code = LEVEL_BY_NAME.get(level_word)
            product_id = get_or_create_product(cur, 'textbook', book_name, level_code, None, unit_price, None)
            po_supplier_groups.setdefault(supplier_name, []).append((product_id, qty, unit_price, shop_price or (unit_price * int(qty))))
            products_added += 1

        for supplier_name, items in po_supplier_groups.items():
            sup_id = get_or_create_supplier(cur, supplier_name)
            total = sum(amt for _,_,_,amt in items)
            cur.execute("""
                insert into purchase_orders (supplier_id, ordered_at, total_amount, notes, source_file)
                values (%s, %s, %s, %s, %s) returning id
            """, (sup_id, datetime.date(2026,1,1), total, f"ESL textbook — {order_label}", ESL_OPENING.name))
            po_id = cur.fetchone()[0]
            pos_added += 1
            for product_id, qty, unit_cost, amount in items:
                cur.execute("""
                    insert into po_items (po_id, product_id, qty, unit_cost, amount)
                    values (%s, %s, %s, %s, %s)
                """, (po_id, product_id, float(qty), unit_cost, amount))
                items_added += 1
                cur.execute("""
                    insert into inventory_movements (product_id, direction, qty, reason,
                        related_po_id, unit_cost, notes)
                    values (%s, 'IN', %s, 'purchase', %s, %s, %s)
                """, (product_id, float(qty), po_id, unit_cost, f"ESL opening — {order_label}"))
                movements_added += 1

    # T-Shirt sheet
    if "T Shirt" in wb.sheetnames:
        ws = wb["T Shirt"]
        rows = list(ws.iter_rows(values_only=True))
        # Find header row with "Size" then size→qty mapping
        ts_items: list[tuple[str, int, int]] = []  # (size, qty, unit_cost)
        for r in rows:
            if not r: continue
            for i in range(len(r) - 1):
                cell = r[i]
                next_cell = r[i + 1] if i + 1 < len(r) else None
                if isinstance(cell, str) and cell.strip().upper() in ("XS","S","M","L","XL","XXL","XXXL","BXL","SXL"):
                    if isinstance(next_cell, (int, float)) and next_cell > 0:
                        ts_items.append((cell.strip().upper(), int(next_cell), 7000))
        if ts_items:
            sup_id = get_or_create_supplier(cur, "Icon Kids (T-Shirt)")
            cur.execute("""
                insert into purchase_orders (supplier_id, ordered_at, total_amount, notes, source_file)
                values (%s, %s, %s, %s, %s) returning id
            """, (sup_id, datetime.date(2026,1,1), sum(q*c for _,q,c in ts_items), "T-Shirt opening order", ESL_OPENING.name))
            po_id = cur.fetchone()[0]; pos_added += 1
            for size, qty, unit in ts_items:
                pid = get_or_create_product(cur, 'tshirt', f"T-Shirt {size}", None, size, unit, 15000)
                cur.execute("insert into po_items (po_id, product_id, qty, unit_cost, amount) values (%s,%s,%s,%s,%s)",
                    (po_id, pid, qty, unit, qty*unit))
                items_added += 1
                cur.execute("""insert into inventory_movements (product_id, direction, qty, reason,
                    related_po_id, unit_cost, notes) values (%s,'IN',%s,'purchase',%s,%s,'T-Shirt opening')""",
                    (pid, qty, po_id, unit))
                movements_added += 1
                products_added += 1

    # ESL Accessories
    if "ESL Accessories" in wb.sheetnames:
        ws = wb["ESL Accessories"]
        rows = list(ws.iter_rows(values_only=True))
        sup_id = get_or_create_supplier(cur, "Misc (ESL Accessories)")
        items_added_acc = 0
        for r in rows:
            if not r or not r[0]: continue
            if isinstance(r[0], (int, float)) and len(r) > 3:
                name = str(r[1]).strip() if r[1] else None
                qty  = r[2] if len(r) > 2 else None
                unit = to_int(r[3]) if len(r) > 3 else 0
                if name and qty and unit:
                    pid = get_or_create_product(cur, 'accessory', name, None, None, unit, None)
                    items_added_acc += 1
                    products_added += 1
        if items_added_acc:
            print(f"    accessories products: {items_added_acc}")
    wb.close()

    conn.commit()
    print(f"  products={products_added}  purchase_orders={pos_added}  po_items={items_added}  movements={movements_added}")

def main():
    conn = psycopg2.connect(URL, connect_timeout=15)
    print("Loading daily income → invoices/payments/ledger…")
    load_daily_income(conn)
    print("\nLoading opening costs → suppliers/products/POs/movements…")
    load_opening_costs(conn)
    conn.close()
    print("\nDone.")

if __name__ == "__main__":
    main()
